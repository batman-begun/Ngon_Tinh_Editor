from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook import Workbook

from omni_novel_editor.config import DEFAULT_OUTPUT_HEADER


def load_workbook_from_bytes(data: bytes) -> Workbook:
    return load_workbook(BytesIO(data))


def save_workbook_to_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def save_workbook_to_path(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def sheet_preview(workbook: Workbook, sheet_name: str, rows: int = 10) -> list[list[Any]]:
    sheet = workbook[sheet_name]
    preview = []
    max_row = min(sheet.max_row, rows)
    max_col = sheet.max_column
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        preview.append(list(row))
    return preview


def column_options(workbook: Workbook, sheet_name: str) -> dict[str, str]:
    sheet = workbook[sheet_name]
    options = {"": "Không chọn"}
    for col_idx in range(1, sheet.max_column + 1):
        letter = get_column_letter(col_idx)
        header = sheet.cell(row=1, column=col_idx).value
        label = f"{letter}"
        if header:
            label += f" — {header}"
        options[letter] = label
    return options


def find_or_create_output_column(workbook: Workbook, sheet_name: str, header: str = DEFAULT_OUTPUT_HEADER) -> str:
    sheet = workbook[sheet_name]
    for col_idx in range(1, sheet.max_column + 1):
        if str(sheet.cell(row=1, column=col_idx).value or "").strip() == header:
            return get_column_letter(col_idx)
    output_idx = sheet.max_column + 1
    sheet.cell(row=1, column=output_idx).value = header
    return get_column_letter(output_idx)


def first_blank_output_row(workbook: Workbook, sheet_name: str, output_col: str, start_row: int = 2) -> int:
    sheet = workbook[sheet_name]
    output_idx = column_index_from_string(output_col)
    for row_idx in range(start_row, sheet.max_row + 1):
        if not str(sheet.cell(row=row_idx, column=output_idx).value or "").strip():
            return row_idx
    return sheet.max_row + 1


def cell_text(workbook: Workbook, sheet_name: str, row: int, col_letter: str | None) -> str:
    if not col_letter:
        return ""
    sheet = workbook[sheet_name]
    value = sheet[f"{col_letter}{row}"].value
    return "" if value is None else str(value)


def set_cell_text(workbook: Workbook, sheet_name: str, row: int, col_letter: str, value: str) -> None:
    workbook[sheet_name][f"{col_letter}{row}"].value = value
